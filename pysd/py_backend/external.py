"""
These classes are a collection of the needed tools to read external data.
The External type objects created by these classes are initialized before
the Stateful objects by functions.Model.initialize.
"""

import re
import warnings

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

import numpy as np
import xarray as xr
import pandas as pd

from . import utils
from .data import Data
from .lookups import Lookups


_SPREADSHEET_EXTS = {'.xls', '.xlsx', '.xlsm', '.xlsb', '.odf', '.ods', '.odt'}


class Excels():
    """
    Class to save the read Excel files and thus avoid double reading
    """
    _Excels, _Excels_opyxl = {}, {}

    @classmethod
    def read(cls, file_name, tab):
        """
        Read the Excel file or return the previously read one
        """
        if file_name.joinpath(tab) in cls._Excels:
            return cls._Excels[file_name.joinpath(tab)]
        else:
            # get the function to read the data based on its extension
            read_kwargs = {}
            ext = file_name.suffix.lower()
            if ext in _SPREADSHEET_EXTS:
                read_func = pd.read_excel
                read_kwargs['sheet_name'] = tab
            elif ext == '.csv':
                read_func = pd.read_csv
                if tab and not tab[0].isalnum():
                    read_kwargs['sep'] = tab
            else:
                read_func = pd.read_table
                if tab and not tab[0].isalnum():
                    read_kwargs['sep'] = tab
            # read the data
            excel = np.array([
                pd.to_numeric(ex, errors='coerce')
                for ex in
                read_func(file_name, header=None, **read_kwargs).values
                ])
            # save data for future retrievals
            cls._Excels[file_name.joinpath(tab)] = excel
            return excel

    @classmethod
    def read_opyxl(cls, file_name):
        """
        Read the Excel file using OpenPyXL or return the previously read one
        """
        if file_name in cls._Excels_opyxl:
            return cls._Excels_opyxl[file_name]
        else:
            excel = load_workbook(file_name, read_only=True, data_only=True)
            cls._Excels_opyxl[file_name] = excel
            return excel

    @classmethod
    def clean(cls):
        """
        Clean the dictionary of read files
        """
        for file in cls._Excels_opyxl.values():
            # close files open directly with openpyxls
            file.close()
            # files open with pandas are automatically closed
        cls._Excels, cls._Excels_opyxl = {}, {}


class External(object):
    """
    Main class of external objects

    Attributes
    ----------
    py_name: str
        The Python name of the object
    missing: str ("warning", "error", "ignore", "keep")
        What to do with missing values. If "warning" (default)
        shows a warning message and interpolates the values.
        If "raise" raises an error. If "ignore" interpolates
        the values without showing anything. If "keep" it will keep
        the missing values, this option may cause the integration to
        fail, but it may be used to check the quality of the data.
    file: str
        File name from which the data is read.
    tab: str
        Tab name from which the data is read. If file type is not a
        spreadsheet this will be used as a separator.

    """
    missing = "warning"

    def __init__(self, py_name):
        self.py_name = py_name
        self.file = None
        self.tab = None

    def __str__(self):
        return self.py_name

    def _get_data_from_file(self, rows, cols):
        """
        Function to read data from excel file using rows and columns

        Parameters
        ----------
        rows: list of len 2
            first row and last row+1 to be read, starting from 0
        cols:  list of len 2
            first col and last col+1 to be read, starting from 0

        Returns
        -------
        data: pandas.DataFrame, pandas.Series or float
            depending on the shape of the requested data

        """
        # read data
        data = Excels.read(
            self.file,
            self.tab)[rows[0]:rows[1], cols[0]:cols[1]].copy()

        shape = data.shape

        # empty cells
        if shape[0] == 0 or shape[1] == 0:
            raise ValueError(
                self.py_name + "\n"
                "The cells are empty.\n"
                + self._file_sheet
            )

        # if it is a single row remove its dimension
        if shape[1] == 1:
            data = data[:, 0]
        if shape[0] == 1:
            data = data[0]
        return data

    def _get_data_from_file_opyxl(self, cellname):
        """
        Function to read data from excel file using cell range name

        Parameters
        ----------
        cellname: str
            the cell range name

        Returns
        -------
        data: numpy.ndarray or float
            depending on the shape of the requested data
        shape: list
            The shape of the data in 2D.

        """
        # read data
        try:
            excel = Excels.read_opyxl(self.file)
        except InvalidFileException:
            raise ValueError(
                self.py_name + "\n"
                f"Cannot read the file '{self.file}'...\n"
                f"It could happen that cell='{cellname}' was "
                "read as a cell range name due to a wrong "
                "definition of cell value"
            )

        # Get global and local cellrange names
        global_cellranges = excel.defined_names
        local_cellranges = None
        # need to lower the sheetnames as Vensim has no case sensitivity
        for sheet in excel.sheetnames:
            if sheet.lower() == self.tab.lower():
                local_cellranges = excel[sheet].defined_names
                break

        if local_cellranges is None:
            # Error if it is not able to get the localSheetId
            raise ValueError(
                self.py_name + "\n"
                "The sheet doesn't exist...\n"
                + self._file_sheet
            )
        try:
            # Search for local and global names
            cellrange = local_cellranges.get(cellname)\
                        or global_cellranges.get(cellname)
            sheet, cells = next(cellrange.destinations)

            assert sheet.lower() == self.tab.lower()
            self.tab = sheet  # case insensitivity in sheet name

            # Get the cells where the cellrange is defined
            cells = re.split(r":|\$", cells)
            cols = [self._col_to_num(cells[1]), None]
            rows = [int(cells[2])-1, None]
            if len(cells) == 3:
                # 0 dim cell range
                cols[1] = cols[0]+1
                rows[1] = rows[0]+1
            else:
                # array or table
                cols[1] = self._col_to_num(cells[4])+1
                rows[1] = int(cells[5])
            # Use pandas to read the data and return its original shape
            return self._get_data_from_file(rows, cols), \
                [rows[1]-rows[0], cols[1]-cols[0]]

        except (AttributeError, AssertionError):
            # key error if the cellrange doesn't exist in the file or sheet
            raise AttributeError(
               self.py_name + "\n"
               f"The cellrange name '{cellname}'\n"
               "Doesn't exist in:\n" + self._file_sheet
               )

    def _get_series_data(self, series_across, series_row_or_col, cell, size):
        """
        Function thar reads series and data from excel file for
        DATA and LOOKUPS.

        Parameters
        ----------
        series_across: "row", "column" or "name"
            The way to read series file.
        series_row_or_col: int or str
            If series_across is "row" the row number where the series data is.
            If series_across is "column" the column name where the series
            data is.
            If series_across is "name" the cell range name where the series
            data is.
        cell:
            If series_across is not "name, the top left cell where the
            data table starts.
            Else the name of the cell range where the data is.
        size:
            The size of the 2nd dimension of the data.

        Returns
        -------
        series, data: ndarray (1D), ndarray(1D/2D)
            The values of the series and data.

        """
        if series_across == "row":
            # Horizontal data (dimension values in a row)

            # get the dimension values
            first_row, first_col = self._split_excel_cell(cell)
            series = self._get_data_from_file(
                rows=[int(series_row_or_col)-1, int(series_row_or_col)],
                cols=[first_col, None])

            # read data
            data = self._get_data_from_file(
                rows=[first_row, first_row + size],
                cols=[first_col, None]).transpose()

        elif series_across == "column":
            # Vertical data (dimension values in a column)

            # get the dimension values
            first_row, first_col = self._split_excel_cell(cell)
            series_col = self._col_to_num(series_row_or_col)
            series = self._get_data_from_file(
                rows=[first_row, None],
                cols=[series_col, series_col+1])

            # read data
            data = self._get_data_from_file(
                rows=[first_row, None],
                cols=[first_col, first_col + size])

        else:
            # get series data
            series, s_shape = self._get_data_from_file_opyxl(series_row_or_col)

            if isinstance(series, float):
                series = np.array([series])

            if s_shape[0] > 1 and s_shape[1] > 1:
                # Error if the lookup/time dimension is 2D
                raise ValueError(
                  self.py_name + "\n"
                  + "Dimension given in:\n"
                  + self._file_sheet
                  + "\tDimension name:"
                  + "\t'{}'\n".format(series_row_or_col)
                  + " is a table and not a vector"
                  )
            elif s_shape[1] != 1:
                transpose = True
            else:
                transpose = False

            # get data
            data, d_shape = self._get_data_from_file_opyxl(cell)

            if isinstance(data, float):
                data = np.array([data])

            if transpose:
                # transpose for horizontal definition of dimension
                data = data.transpose()
                d_shape = d_shape[1], d_shape[0]

            if d_shape[0] != len(series):
                raise ValueError(
                  self.py_name + "\n"
                  + "Dimension and data given in:\n"
                  + self._file_sheet
                  + "\tDimension name:\t'{}'\n".format(series_row_or_col)
                  + "\tData name:\t'{}'\n".format(cell)
                  + " don't have the same length in the 1st dimension"
                  )

            if d_shape[1] != size:
                # Given coordinates length is different than
                # the lentgh of 2nd dimension
                raise ValueError(
                  self.py_name + "\n"
                  + "Data given in:\n"
                  + self._file_sheet
                  + "\tData name:\t'{}'\n".format(cell)
                  + " has not the same size as the given coordinates"
                  )

        return series, data

    def _resolve_file(self, root):
        """
        Resolve input file path. Joining the file with the root and
        checking if it exists.

        Parameters
        ----------
        root: pathlib.Path or str
            The root path to the model file.

        Returns
        -------
        None

        """
        if str(self.file)[0] == '?':
            # TODO add an option to include indirect references
            raise ValueError(
                self.py_name + "\n"
                + f"Indirect reference to file: '{self.file}'")

        # Join path and resolve it to better print error messages
        self.file = root.joinpath(self.file).resolve()

        if not self.file.is_file():
            raise FileNotFoundError(
                self.py_name + "\n"
                + "File '%s' not found." % self.file)

    def _initialize_data(self, element_type):
        """
        Initialize one element of DATA or LOOKUPS

        Parameters
        ----------
        element_type: str
            "lookup" for LOOKUPS, "data" for data.

        Returns
        -------
        data: xarray.DataArray
            Dataarray with the time or interpolation dimension
            as first dimension.

        """
        self._resolve_file(root=self.root)
        series_across = self._series_selector(self.x_row_or_col, self.cell)
        size = utils.compute_shape(self.coords, reshape_len=1,
                                   py_name=self.py_name)[0]

        series, data = self._get_series_data(
            series_across=series_across,
            series_row_or_col=self.x_row_or_col,
            cell=self.cell, size=size
        )

        # remove nan or missing values from dimension
        if series_across != "name":
            # Remove last nans only if the method is to read by row or col
            i = 0
            try:
                while np.isnan(series[i-1]):
                    i -= 1
            except IndexError:
                # series has len 0
                raise ValueError(
                    self.py_name + "\n"
                    + "Dimension given in:\n"
                    + self._file_sheet
                    + "\t{}:\t'{}'\n".format(series_across, self.x_row_or_col)
                    + " has length 0"
                    )

            if i != 0:
                series = series[:i]
                data = data[:i]

        # warning/error if missing data in the series
        if any(np.isnan(series)) and self.missing != "keep":
            valid_values = ~np.isnan(series)
            series = series[valid_values]
            data = data[valid_values]
            if all(np.isnan(series)):
                raise ValueError(
                    self.py_name + "\n"
                    + "Dimension given in:\n"
                    + self._file_sheet
                    + "\t{}:\t'{}'\n".format(series_across, self.x_row_or_col)
                    + " has length 0"
                    )
            if self.missing == "warning":
                warnings.warn(
                  self.py_name + "\n"
                  + "Dimension value missing or non-valid in:\n"
                  + self._file_sheet
                  + "\t{}:\t'{}'\n".format(series_across, self.x_row_or_col)
                  + " the corresponding data value(s) to the "
                  + "missing/non-valid value(s) will be ignored\n\n"
                  )
            elif self.missing == "raise":
                raise ValueError(
                  self.py_name + "\n"
                  + "Dimension value missing or non-valid in:\n"
                  + self._file_sheet
                  + "\t{}:\t'{}'\n".format(series_across, self.x_row_or_col)
                  )

        # reorder data with increasing series
        if not np.all(np.diff(series) > 0) and self.missing != "keep":
            order = np.argsort(series)
            series = series[order]
            data = data[order]
            # Check if the lookup/time dimension is well defined
            if np.any(np.diff(series) == 0):
                raise ValueError(self.py_name + "\n"
                                 + "Dimension given in:\n"
                                 + self._file_sheet
                                 + "\t{}:\t'{}'\n".format(
                                    series_across, self.x_row_or_col)
                                 + " has repeated values")

        # Check for missing values in data
        if np.any(np.isnan(data)) and self.missing != "keep":
            if series_across == "name":
                cell_type = "Cellrange"
            else:
                cell_type = "Reference cell"

            if self.missing == "warning":
                # Fill missing values with the chosen interpolation method
                # what Vensim does during running for DATA
                if self.interp != "raw":
                    interpolate_message =\
                        " the corresponding value will be filled "\
                        + "with the interpolation method of the object."
                else:
                    interpolate_message = ""

                warnings.warn(
                    self.py_name + "\n"
                    + "Data value missing or non-valid in:\n"
                    + self._file_sheet
                    + "\t{}:\t'{}'\n".format(cell_type, self.cell)
                    + interpolate_message + "\n\n"
                    )
            elif self.missing == "raise":
                raise ValueError(
                    self.py_name + "\n"
                    + "Data value missing or non-valid in:\n"
                    + self._file_sheet
                    + "\t{}:\t'{}'\n".format(cell_type, self.cell)
                    )
            # fill values
            if self.interp != "raw":
                self._fill_missing(series, data)

        # reshape the data to fit in the xarray.DataArray
        reshape_dims = tuple([len(series)] + utils.compute_shape(self.coords))
        data = self._reshape(data, reshape_dims)

        if element_type == "lookup":
            dim_name = "lookup_dim"
        else:
            dim_name = "time"

        data = xr.DataArray(
            data=data,
            coords={dim_name: series, **self.coords},
            dims=[dim_name] + list(self.coords)
        )

        return data

    def _fill_missing(self, series, data):
        """
        Fills missing values in excel read data. Mutates the values in data.

        Parameters
        ----------
        series:
          the time series without missing values
        data:
          the data with missing values

        Returns
        -------
        None
        """
        # if data is 2dims we need to interpolate
        datanan = np.isnan(data)
        keeping_nan = False
        if len(data.shape) == 1:
            if not np.all(datanan):
                data[datanan] = self._interpolate_missing(
                    series[datanan],
                    series[~datanan],
                    data[~datanan])
            else:
                keeping_nan = True
        else:
            for i, nanlist in enumerate(list(datanan.transpose())):
                if not np.all(nanlist):
                    data[nanlist, i] = self._interpolate_missing(
                        series[nanlist],
                        series[~nanlist],
                        data[~nanlist][:, i])
                else:
                    keeping_nan = True

        if keeping_nan:
            warnings.warn(
                "Not able to interpolate some values..."
                " keeping them as missing.\n")

    def _interpolate_missing(self, x, xr, yr):
        """
        Interpolates a list of missing values from _fill_missing

        Parameters
        ----------
        x:
          list of missing values interpolate
        xr:
          non-missing x values
        yr:
          non-missing y values

        Returns
        -------
        y:
          Result after interpolating x with self.interp method

        """
        y = np.empty_like(x, dtype=float)
        for i, value in enumerate(x):
            if value >= xr[-1]:
                y[i] = yr[-1]
            elif value <= xr[0]:
                y[i] = yr[0]
            elif self.interp == 'look_forward':
                y[i] = yr[xr >= value][0]
            elif self.interp == 'hold_backward':
                y[i] = yr[xr <= value][-1]
            else:
                y[i] = np.interp(value, xr, yr)
        return y

    @property
    def _file_sheet(self):
        """
        Returns file and sheet name in a string
        """
        return "\tFile name:\t'{}'\n".format(self.file)\
               + "\tSheet name:\t'{}'\n".format(self.tab)

    @staticmethod
    def _col_to_num(col):
        """
        Transforms the column name to int

        Parameters
        ----------
        col: str
          Column name

        Returns
        -------
        int
          Column number
        """
        if len(col) == 1:
            return ord(col.upper()) - ord('A')
        elif len(col) == 2:
            left = ord(col[0].upper()) - ord('A') + 1
            right = ord(col[1].upper()) - ord('A')
            return left * (ord('Z')-ord('A')+1) + right
        else:
            left = ord(col[0].upper()) - ord('A') + 1
            center = ord(col[1].upper()) - ord('A') + 1
            right = ord(col[2].upper()) - ord('A')
            return left * ((ord('Z')-ord('A')+1)**2)\
                + center * (ord('Z')-ord('A')+1)\
                + right

    def _split_excel_cell(self, cell):
        """
        Splits a cell value given in a string.
        Returns None for non-valid cell formats.

        Parameters
        ----------
        cell: str
          Cell like string, such as "A1", "b16", "AC19"...
          If it is not a cell like string will return None.

        Returns
        -------
        row number, column number: int, int
          If the cell input is valid. Both numbers are given in Python
          enumeration, i.e., first row and first column are 0.

        """
        split = re.findall(r'\d+|\D+', cell)
        try:
            # check that we only have two values [column, row]
            assert len(split) == 2
            # check that the column name has no special characters
            assert not re.compile('[^a-zA-Z]+').search(split[0])
            # check that row number is not 0
            assert int(split[1]) != 0
            # the column name has as maximum 3 letters
            assert len(split[0]) <= 3
            return int(split[1])-1, self._col_to_num(split[0])
        except AssertionError:
            return

    @staticmethod
    def _reshape(data, dims):
        """
        Reshapes an pandas.DataFrame, pandas.Series, xarray.DataArray
        or np.ndarray in the given dimensions.

        Parameters
        ----------
        data: xarray.DataArray/numpy.ndarray
          Data to be reshaped
        dims: tuple
          The dimensions to reshape.

        Returns
        -------
        numpy.ndarray
          reshaped array
        """
        if isinstance(data, (float, int)):
            data = np.array(data)
        elif isinstance(data, xr.DataArray):
            data = data.values

        return data.reshape(dims)

    def _series_selector(self, x_row_or_col, cell):
        """
        Selects if a series data (DATA/LOOKUPS), should be read by columns,
        rows or cellrange name.
        Based on the input format of x_row_or_col and cell.
        The format of the 2 variables must be consistent.

        Parameters
        ----------
        x_row_or_col: str
          String of a number if series is given in a row, letter if series is
          given in a column or name if the series is given by cellrange name.
        cell: str
          Cell identificator, such as "A1", or name if the data is given
          by cellrange name.

        Returns
        -------
        series_across: str
          "row" if series is given in a row
          "column" if series is given in a column
          "name" if series and data are given by range name

        """
        try:
            # if x_row_or_col is numeric the series must be a row
            int(x_row_or_col)
            return "row"

        except ValueError:
            if self._split_excel_cell(cell):
                # if the cell can be splitted means that the format is
                # "A1" like then the series must be a column
                return "column"
            else:
                return "name"


class ExtData(External, Data):
    """
    Class for Vensim GET XLS DATA/GET DIRECT DATA
    """

    def __init__(self, file_name, tab, time_row_or_col, cell,
                 interp, coords, root, final_coords, py_name):
        super().__init__(py_name)
        self.files = [file_name]
        self.tabs = [tab]
        self.time_row_or_cols = [time_row_or_col]
        self.cells = [cell]
        self.coordss = [coords]
        self.root = root
        self.final_coords = final_coords
        self.interp = interp or "interpolate"
        self.is_float = not bool(coords)

        # check if the interpolation method is valid
        if self.interp not in ["interpolate", "raw",
                               "look_forward", "hold_backward"]:
            raise ValueError(self.py_name + "\n"
                             + " The interpolation method (interp) must be "
                             + "'raw', 'interpolate', "
                             + "'look_forward' or 'hold_backward'")

    def add(self, file_name, tab, time_row_or_col, cell, interp, coords):
        """
        Add information to retrieve new dimension in an already declared object
        """
        self.files.append(file_name)
        self.tabs.append(tab)
        self.time_row_or_cols.append(time_row_or_col)
        self.cells.append(cell)
        self.coordss.append(coords)

        interp = interp or "interpolate"

        if interp.replace(" ", "_") != self.interp:
            raise ValueError(self.py_name + "\n"
                             + "Error matching interpolation method with "
                             + "previously defined one")

        if list(coords) != list(self.coordss[0]):
            raise ValueError(self.py_name + "\n"
                             + "Error matching dimensions with previous data")

    def initialize(self):
        """
        Initialize all elements and create the self.data xarray.DataArray
        """
        if not self.coordss[0]:
            # Just load one value (no add)
            for self.file, self.tab, self.x_row_or_col, \
                self.cell, self.coords\
                in zip(self.files, self.tabs, self.time_row_or_cols,
                       self.cells, self.coordss):
                self.data = self._initialize_data("data")
        else:
            # Load in several lines (add)
            self.data = xr.DataArray(
                np.nan, self.final_coords, list(self.final_coords))

            for self.file, self.tab, self.x_row_or_col, \
                self.cell, self.coords\
                in zip(self.files, self.tabs, self.time_row_or_cols,
                       self.cells, self.coordss):
                values = self._initialize_data("data")

                coords = {"time": values.coords["time"].values, **self.coords}
                if "time" not in self.data.dims:
                    self.data = self.data.expand_dims(
                        {"time": coords["time"]}, axis=0).copy()

                self.data.loc[coords] = values.values

        # set what to return when raw
        if self.final_coords:
            self.nan = xr.DataArray(
                np.nan, self.final_coords, list(self.final_coords))
        else:
            self.nan = np.nan


class ExtLookup(External, Lookups):
    """
    Class for Vensim GET XLS LOOKUPS/GET DIRECT LOOKUPS
    """

    def __init__(self, file_name, tab, x_row_or_col, cell, coords,
                 root, final_coords, py_name):
        super().__init__(py_name)
        self.files = [file_name]
        self.tabs = [tab]
        self.x_row_or_cols = [x_row_or_col]
        self.cells = [cell]
        self.coordss = [coords]
        self.root = root
        self.final_coords = final_coords
        self.interp = "interpolate"
        self.is_float = not bool(coords)

    def add(self, file_name, tab, x_row_or_col, cell, coords):
        """
        Add information to retrieve new dimension in an already declared object
        """
        self.files.append(file_name)
        self.tabs.append(tab)
        self.x_row_or_cols.append(x_row_or_col)
        self.cells.append(cell)
        self.coordss.append(coords)

        if list(coords) != list(self.coordss[0]):
            raise ValueError(self.py_name + "\n"
                             + "Error matching dimensions with previous data")

    def initialize(self):
        """
        Initialize all elements and create the self.data xarray.DataArray
        """
        if not self.coordss[0]:
            # Just loag one value (no add)
            for self.file, self.tab, self.x_row_or_col, \
                self.cell, self.coords\
                in zip(self.files, self.tabs, self.x_row_or_cols,
                       self.cells, self.coordss):
                self.data = self._initialize_data("lookup")
        else:
            # Load in several lines (add)
            self.data = xr.DataArray(
                np.nan, self.final_coords, list(self.final_coords))

            for self.file, self.tab, self.x_row_or_col, \
                self.cell, self.coords\
                in zip(self.files, self.tabs, self.x_row_or_cols,
                       self.cells, self.coordss):
                values = self._initialize_data("lookup")

                coords = {
                    "lookup_dim": values.coords["lookup_dim"].values,
                    **self.coords
                }
                if "lookup_dim" not in self.data.dims:
                    self.data = self.data.expand_dims(
                        {"lookup_dim": coords["lookup_dim"]}, axis=0).copy()

                self.data.loc[coords] = values.values


class ExtConstant(External):
    """
    Class for Vensim GET XLS CONSTANTS/GET DIRECT CONSTANTS
    """

    def __init__(self, file_name, tab, cell, coords,
                 root, final_coords, py_name):
        super().__init__(py_name)
        self.files = [file_name]
        self.tabs = [tab]
        self.transposes = [
            cell[-1] == '*' and np.prod(utils.compute_shape(coords)) > 1]
        self.cells = [cell.strip('*')]
        self.coordss = [coords]
        self.root = root
        self.final_coords = final_coords

    def add(self, file_name, tab, cell, coords):
        """
        Add information to retrieve new dimension in an already declared object
        """
        self.files.append(file_name)
        self.tabs.append(tab)
        self.transposes.append(
            cell[-1] == '*' and np.prod(utils.compute_shape(coords)) > 1)
        self.cells.append(cell.strip('*'))
        self.coordss.append(coords)

        if list(coords) != list(self.coordss[0]):
            raise ValueError(self.py_name + "\n"
                             + "Error matching dimensions with previous data")

    def initialize(self):
        """
        Initialize all elements and create the self.data xarray.DataArray
        """
        if not self.coordss[0]:
            # Just loag one value (no add)
            for self.file, self.tab, self.transpose, self.cell, self.coords\
                in zip(self.files, self.tabs, self.transposes,
                       self.cells, self.coordss):
                self.data = self._initialize()
        else:
            # Load in several lines (add)

            self.data = xr.DataArray(
                np.nan, self.final_coords, list(self.final_coords))

            for self.file, self.tab, self.transpose, self.cell, self.coords\
                in zip(self.files, self.tabs, self.transposes,
                       self.cells, self.coordss):
                self.data.loc[self.coords] = self._initialize().values

    def _initialize(self):
        """
        Initialize one element
        """
        self._resolve_file(root=self.root)
        split = self._split_excel_cell(self.cell)
        if split:
            data_across = "cell"
            cell = split
        else:
            data_across = "name"
            cell = self.cell

        shape = utils.compute_shape(self.coords, reshape_len=2,
                                    py_name=self.py_name)

        if self.transpose:
            shape.reverse()

        data = self._get_constant_data(data_across, cell, shape)

        if self.transpose:
            data = data.transpose()

        if np.any(np.isnan(data)):
            # nan values in data
            if data_across == "name":
                cell_type = "Cellrange"
            else:
                cell_type = "Reference cell"

            if self.missing == "warning":
                warnings.warn(
                    self.py_name + "\n"
                    + "Constant value missing or non-valid in:\n"
                    + self._file_sheet
                    + "\t{}:\t'{}'\n".format(cell_type, self.cell)
                    )
            elif self.missing == "raise":
                raise ValueError(
                    self.py_name + "\n"
                    + "Constant value missing or non-valid in:\n"
                    + self._file_sheet
                    + "\t{}:\t'{}'\n".format(cell_type, self.cell)
                    )

        # Create only an xarray if the data is not 0 dimensional
        if len(self.coords) > 0:
            reshape_dims = tuple(utils.compute_shape(self.coords))
            data = self._reshape(data, reshape_dims)
            data = xr.DataArray(
                data=data, coords=self.coords, dims=list(self.coords)
            )

        return data

    def _get_constant_data(self, data_across, cell, shape):
        """
        Function thar reads data from excel file for CONSTANT

        Parameters
        ----------
        data_across: "cell" or "name"
            The way to read data file.
        cell: int or str
            If data_across is "cell" the lefttop split cell value where
            the data is.
            If data_across is "name" the cell range name where the data is.
        shape: list
            The shape of the data in 2D.

        Returns
        -------
        data: float/ndarray(1D/2D)
            The values of the data.

        """
        if data_across == "cell":
            # read data from topleft cell name using pandas
            start_row, start_col = cell

            return self._get_data_from_file(
                rows=[start_row, start_row + shape[0]],
                cols=[start_col, start_col + shape[1]])

        else:
            # read data from cell range name using OpenPyXL
            data, xl_shape = self._get_data_from_file_opyxl(cell)
            if shape != xl_shape:
                raise ValueError(self.py_name + "\n"
                                 + "Data given in:\n"
                                 + self._file_sheet
                                 + "\tData name:\t{}\n".format(cell)
                                 + " has not the same shape as the"
                                 + " given coordinates")

            return data

    def __call__(self):
        return self.data


class ExtSubscript(External):
    """
    Class for Vensim GET XLS SUBSCRIPT/GET DIRECT SUBSCRIPT
    """

    def __init__(self, file_name, tab, firstcell, lastcell, prefix, root):
        super().__init__("Hardcoded external subscript")
        self.file = file_name
        self.tab = tab
        self.prefix = prefix
        self._resolve_file(root=root)
        split = self._split_excel_cell(firstcell)
        if split:
            subs = self.get_subscripts_cell(*split, lastcell)
        else:
            subs = self.get_subscripts_name(firstcell)

        self.subscript = [
            self.prefix + str(d) for d in subs.flatten()
            if self._not_nan(d)
            ]

    def get_subscripts_cell(self, row_first, col_first, lastcell):
        """Get subscripts from common cell definition"""
        if not lastcell:
            row_last, col_last = None, None
        else:
            split = self._split_excel_cell(lastcell)
            if split:
                # last cell is col and row
                row_last, col_last = split
            elif lastcell.isdigit():
                # last cell is row number only
                row_last = int(lastcell)-1
                col_last = None
            else:
                # last cell is a col value only
                row_last = None
                col_last = self._col_to_num(lastcell)

        # update read keywargs for rows and columns to read
        read_kwargs = {}
        if row_last is not None:
            read_kwargs['nrows'] = row_last-row_first+1
        if col_last is not None:
            read_kwargs['usecols'] = np.arange(col_first, col_last+1)

        # get the function to read the data based on its extension
        ext = self.file.suffix.lower()
        if ext in _SPREADSHEET_EXTS:
            read_func = pd.read_excel
            read_kwargs['sheet_name'] = self.tab
        elif ext == '.csv':
            read_func = pd.read_csv
            if self.tab and not self.tab[0].isalnum():
                read_kwargs['sep'] = self.tab
        else:
            read_func = pd.read_table
            if self.tab and not self.tab[0].isalnum():
                read_kwargs['sep'] = self.tab

        # read the data
        data = read_func(
            self.file,
            skiprows=row_first,
            dtype=object,
            header=None,
            **read_kwargs
            ).values

        # skip columns if usecols couldn't be used
        if col_last is None:
            data = data[:, col_first:]

        return data

    def get_subscripts_name(self, cellname):
        """Get subscripts from cell range name definition"""
        try:
            excel = load_workbook(self.file, read_only=True, data_only=True)
        except InvalidFileException:
            raise ValueError(
                self.py_name + "\n"
                f"Cannot read the file '{self.file}'...\n"
                f"It could happen that firstcell='{cellname}' was "
                "read as a cell range name due to a wrong definition "
                "of cell value"
            )
        global_cellranges = excel.defined_names
        local_cellranges = None
        # need to lower the sheetnames as Vensim has no case sensitivity
        for sheet in excel.sheetnames:
            if sheet.lower() == self.tab.lower():
                local_cellranges = excel[sheet].defined_names
                break

        if local_cellranges is None:
            # Error if it is not able to get the localSheetId
            raise ValueError(
                self.py_name + "\n"
                "The sheet doesn't exist...\n"
                + self._file_sheet
            )
        try:
            # Search for local and global names
            cellrange = local_cellranges.get(cellname)\
                        or global_cellranges.get(cellname)
            sheet, cells = next(cellrange.destinations)

            assert sheet.lower() == self.tab.lower()
            self.tab = sheet  # case insensitivity in sheet name

            # Get the cells where the cellrange is defined
            first_cell, last_cell = cells.replace("$", '').split(":")
        except (AttributeError, AssertionError):
            # key error if the cellrange doesn't exist in the file or sheet
            raise AttributeError(
               self.py_name + "\n"
               f"The cellrange name '{cellname}'\n"
               "Doesn't exist in:\n" + self._file_sheet
               )
        else:
            return self.get_subscripts_cell(
                *self._split_excel_cell(first_cell), last_cell)

    @staticmethod
    def _not_nan(value):
        """Check if a value is not nan"""
        if isinstance(value, str):
            return True
        return not np.isnan(value)
